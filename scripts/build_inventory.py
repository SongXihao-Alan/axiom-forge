#!/usr/bin/env python3
"""Build downloads_inventory.xlsx — categorize every file in ~/Downloads/."""

import os
from pathlib import Path
from datetime import datetime
from collections import Counter
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DL = Path("/Users/alan/Downloads")
AF = DL / "axiom-finder"

def fmt_mtime(p):
    try:
        return datetime.fromtimestamp(p.stat().st_mtime).strftime("%Y-%m-%d")
    except Exception:
        return "?"

def fmt_size(p):
    try:
        s = p.stat().st_size
        if s > 1024 * 1024:
            return f"{s/1024/1024:.1f} MB"
        if s > 1024:
            return f"{s/1024:.1f} KB"
        return f"{s} B"
    except Exception:
        return "?"

rows = []

def add(zone, project, file_rel, abs_p, status, purpose):
    if isinstance(abs_p, str):
        abs_p = Path(abs_p)
    rel = str(abs_p.relative_to(DL)) if abs_p.exists() else file_rel
    rows.append({
        "zone": zone,
        "project": project,
        "file": abs_p.name if abs_p.exists() else file_rel.split("/")[-1],
        "rel_path": rel,
        "size": fmt_size(abs_p) if abs_p.exists() else "—",
        "mtime": fmt_mtime(abs_p) if abs_p.exists() else "—",
        "status": status,
        "purpose": purpose,
    })

# ── Downloads 根 ──
add("Downloads 根目录", "(临时)", "~$*.docx 共 10 个", DL / "~$博士后人才支持项目申请表.docx",
    "临时", "Word 临时锁文件 (~$,162B),Word 关闭后会自动删除;共 10 个类似文件分散在根目录")
add("Downloads 根目录", "(参考)", "The Axiomatics of Economic Design.pdf", DL / "The Axiomatics of Economic Design.pdf",
    "历史/参考", "Thomson 等不可能性定理经典著作 PDF,4.0 MB,作为 Phase D Thomson 2023 形式化的原始参考")
add("Downloads 根目录", "(工具)", "downloads_inventory.xlsx", DL / "downloads_inventory.xlsx",
    "独立项目", "上轮生成的 Downloads 全量文件清单 Excel,会持续更新")

# ── research applied/ — 另一个科研项目 ──
add("research applied/", "信贷市场研究", "research_proposal.md", DL / "research applied" / "research_proposal.md",
    "独立项目", "信贷市场信息异质性、银行竞争与征信共享政策研究 proposal(中国情境)")
add("research applied/", "信贷市场研究", "3文献精读报告.md", DL / "research applied" / "3文献精读报告.md",
    "独立项目", "意大利汽车保险/信贷市场 3 篇核心文献精读")
add("research applied/", "信贷市场研究", "Asymmetric Information and Imperfect Competition in Lending Markets.pdf",
    DL / "research applied" / "Asymmetric Information and Imperfect Competition in Lending Markets.pdf",
    "独立项目", "信贷市场信息不对称文献")
add("research applied/", "信贷市场研究", "Competing under Information Heterogeneity Evidence from Auto Insurance.pdf",
    DL / "research applied" / "Competing under Information Heterogeneity Evidence from Auto Insurance.pdf",
    "独立项目", "汽车保险信息异质性文献")
add("research applied/", "信贷市场研究", "INFORMATION SPAN IN CREDIT MARKET COMPETITION.pdf",
    DL / "research applied" / "INFORMATION SPAN IN CREDIT MARKET COMPETITION.pdf",
    "独立项目", "信贷市场竞争信息跨度文献")
add("research applied/", "信贷市场研究", "变量.xlsx", DL / "research applied" / "变量.xlsx",
    "独立项目", "研究变量定义表 Excel")
add("research applied/", "SHAP 实验", "extended_shap_experiment_fixed_v2.py",
    DL / "research applied" / "extended_shap_experiment_fixed_v2.py",
    "独立项目", "Colab 导出的 SHAP 缺失场景实验 notebook 源码")
add("research applied/", "SHAP 实验", "shap_conditional_analysis.py",
    DL / "research applied" / "shap_conditional_analysis.py",
    "独立项目", "SHAP 条件分析 Python 脚本(独立运行)")
add("research applied/", "SHAP 实验", "SHAP_缺失场景分析报告.md",
    DL / "research applied" / "SHAP_缺失场景分析报告.md",
    "独立项目", "SHAP b 缺失场景实验分析报告(2026-06-23)")

# ── axiom-finder/ 主项目 ──
# 根
add("axiom-finder/ 根目录", "项目元数据", "README.md", AF / "README.md",
    "活跃", "项目主说明 v0.3-alpha")
add("axiom-finder/ 根目录", "项目元数据", "CONTRIBUTING.md", AF / "CONTRIBUTING.md",
    "活跃", "贡献指南")
add("axiom-finder/ 根目录", "项目元数据", "LICENSE", AF / "LICENSE",
    "活跃", "开源许可证")
add("axiom-finder/ 根目录", "项目元数据", "requirements.txt", AF / "requirements.txt",
    "活跃", "Python 依赖清单")
add("axiom-finder/ 根目录", "项目元数据", "SHIP_TO_GITHUB.sh", AF / "SHIP_TO_GITHUB.sh",
    "活跃", "推送到 GitHub 的发布脚本")
add("axiom-finder/ 根目录", "CLI 入口", "axiom-forge", AF / "axiom-forge",
    "活跃", "顶层 shell 包装脚本,转发 lane-b/lane-c/feedback/compare-versions 到 kb_query 或 lane_b_evaluator")
add("axiom-finder/ 根目录", "Web API", "web_api.py", AF / "web_api.py",
    "活跃", "FastAPI 后端入口,SPA UI 的接口")
add("axiom-finder/ 根目录", "Web 前端", "static/", AF / "static",
    "活跃", "Web SPA UI 静态资源(被 web_api.py 提供)")

# PLAN/
for fname in ["01_master_plan", "03_lane_B_evaluator", "04_lane_C_statistics",
            "05_lane_D_paper", "06_kb_inventory", "07_engineering_flow",
            "08_git_workflow", "09_lean_integration", "10_glossary",
            "11_implementation_checklist", "12_risk_register"]:
    add("axiom-finder/PLAN/", "规划", f"{fname}.md", AF / "PLAN" / f"{fname}.md",
        "活跃", "规划文档主版(MD)")

# kb/
add("axiom-finder/knowledge-base/", "KB Schema", "SCHEMA.md", AF / "knowledge-base" / "SCHEMA.md",
    "活跃", "KB 节点 JSON schema 规范")
add("axiom-finder/knowledge-base/", "KB Schema", "UPDATE_MECHANISM.md", AF / "knowledge-base" / "UPDATE_MECHANISM.md",
    "活跃", "KB 节点更新机制说明")
add("axiom-finder/knowledge-base/", "KB Schema", "REPRODUCTION", AF / "knowledge-base" / "REPRODUCTION",
    "活跃", "KB 重建复现步骤(目录)")
add("axiom-finder/knowledge-base/", "KB Schema", "kb_llm.py", AF / "knowledge-base" / "kb_llm.py",
    "活跃", "KB LLM 调用封装")
add("axiom-finder/knowledge-base/", "KB Schema", "kb_query.py", AF / "knowledge-base" / "kb_query.py",
    "活跃", "13 命令 CLI 主入口(Lane C 查询/feedback/compare-versions)")
add("axiom-finder/knowledge-base/", "KB Schema", "generate_nodes.py", AF / "knowledge-base" / "generate_nodes.py",
    "活跃", "节点生成工具")

kb_dirs = {
    "axioms": "8 个真 axiom 节点",
    "theorems": "10 个定理节点",
    "assumptions": "2 个假设节点",
    "value_anchors": "33 个价值锚节点",
    "diktats": "12 个 diktat 节点",
    "tradeoffs": "4 个权衡节点",
    "scenarios": "6 个场景节点",
    "relations": "关系节点(目录)",
    "literature": "10 个 literature 节点",
    "compressed_views": "压缩视图节点",
    "training_signals": "训练信号节点",
}
for d, desc in kb_dirs.items():
    add("axiom-finder/knowledge-base/nodes/", "KB 数据", f"{d}/", AF / "knowledge-base" / "nodes" / d,
        "活跃", desc)

add("axiom-finder/knowledge-base/nodes/", "KB 数据", "relations.json", AF / "knowledge-base" / "nodes" / "relations.json",
    "活跃", "节点间关系索引")

# kb/ingest/
add("axiom-finder/knowledge-base/ingest/", "KB Ingest", "build_seeds.py", AF / "knowledge-base" / "ingest" / "build_seeds.py",
    "活跃", "种子数据构建")
add("axiom-finder/knowledge-base/ingest/", "KB Ingest", "extract_nodes.py", AF / "knowledge-base" / "ingest" / "extract_nodes.py",
    "活跃", "节点抽取主脚本")
add("axiom-finder/knowledge-base/ingest/", "KB Ingest", "literature_fetcher.py", AF / "knowledge-base" / "ingest" / "literature_fetcher.py",
    "活跃", "Literature API 抓取")
add("axiom-finder/knowledge-base/ingest/", "KB Ingest", "paper_pipeline.py", AF / "knowledge-base" / "ingest" / "paper_pipeline.py",
    "活跃", "论文→节点流水线")
add("axiom-finder/knowledge-base/ingest/", "KB Ingest", "proof_checker.py", AF / "knowledge-base" / "ingest" / "proof_checker.py",
    "活跃", "Lean 证明检查器代理(Phase C)")
add("axiom-finder/knowledge-base/ingest/", "Phase 2 Pipeline 实现", "z3_verify.py", AF / "knowledge-base" / "ingest" / "z3_verify.py",
    "活跃", "Z3 3-tier 验证(Tier A regex / Tier B direct / Tier C LLM→SMT-LIB2);dataclass 驱动;anthropic Claude 后端")
add("axiom-finder/knowledge-base/ingest/", "Phase 2 Pipeline 实现", "discover.py", AF / "knowledge-base" / "ingest" / "discover.py",
    "活跃", "Step D: NL 公理候选发现(call_1_discover,anthropic + instructor)")
add("axiom-finder/knowledge-base/ingest/", "Phase 2 Pipeline 实现", "formalize.py", AF / "knowledge-base" / "ingest" / "formalize.py",
    "活跃", "Step F: 结构化形式化 + SMT-LIB2 片段(call_2_formalize)")
add("axiom-finder/knowledge-base/ingest/", "Phase 2 Pipeline 实现", "backtranslate.py", AF / "knowledge-base" / "ingest" / "backtranslate.py",
    "活跃", "Step 一致性闸:从形式化重建 NL(call_3_backtranslate,sklearn cosine)")
add("axiom-finder/knowledge-base/ingest/", "Phase 2 Pipeline 实现", "pipeline.py", AF / "knowledge-base" / "ingest" / "pipeline.py",
    "活跃", "Phase 2 编排器(D→F→backtranslate→Z3,写 AxiomRecord 到 JSONL)")
add("axiom-finder/knowledge-base/", "Phase 2 中文文档", "SPEC_cn.md", AF / "knowledge-base" / "SPEC_cn.md",
    "活跃", "axiom-forge v0.3 第二步 & 第三步 中文技术文档(从 docx 转换)")
add("axiom-finder/knowledge-base/_archive/", "Phase 2 历史", "z3_verify_v0.py", AF / "knowledge-base" / "_archive" / "z3_verify_v0.py",
    "历史", "v0.3 旧版 z3_verify.py(Lane B 用的精简版,388 行,函数式;被新版 dataclass 版替代)")
add("axiom-finder/knowledge-base/ingest/", "KB Ingest", "lane_b_evaluator.py", AF / "knowledge-base" / "ingest" / "lane_b_evaluator.py",
    "活跃", "Lane B 评估器(5-dim rubric + 3-Layer + Cohen's kappa)")
add("axiom-finder/knowledge-base/ingest/", "Lane B prompts", "lane_b_prompts/v1.md", AF / "knowledge-base" / "ingest" / "lane_b_prompts" / "v1.md",
    "历史", "v1 prompt(67 行),初版,被 v2 取代")
add("axiom-finder/knowledge-base/ingest/", "Lane B prompts", "lane_b_prompts/v2.md", AF / "knowledge-base" / "ingest" / "lane_b_prompts" / "v2.md",
    "历史", "v2 prompt(87 行),v1→v2 中间版")
add("axiom-finder/knowledge-base/ingest/", "Lane B prompts", "lane_b_prompts/v3.md", AF / "knowledge-base" / "ingest" / "lane_b_prompts" / "v3.md",
    "活跃", "v3 prompt(155 行),当前默认,自检 internal_consistency QWK=0.74,clarity=0.72")

# agents/
add("axiom-finder/agents/", "Agents", "gap_finder.py", AF / "agents" / "gap_finder.py",
    "活跃", "候选 axiom 发现代理(Lane B 前置)")
add("axiom-finder/agents/", "Agents", "literature_agent.py", AF / "agents" / "literature_agent.py",
    "活跃", "Literature agent(论文→结构化节点)")
add("axiom-finder/agents/", "Agents", "llm.py", AF / "agents" / "llm.py",
    "活跃", "Agents 共用 LLM 客户端")
add("axiom-finder/agents/", "Agents", "reality_agent.py", AF / "agents" / "reality_agent.py",
    "活跃", "Reality agent(实战测试代理)")
add("axiom-finder/agents/", "Agents", "gap_finder_prompts/", AF / "agents" / "gap_finder_prompts",
    "活跃", "gap_finder prompt 模板目录")

# paper/
add("axiom-finder/paper/", "Lane A", "data/gold.json", AF / "paper" / "data" / "gold.json",
    "活跃", "104 项黄金标注集(74 真 axiom + 30 distractor)")
add("axiom-finder/paper/", "Lane A", "data/gold_dual_annotator.json", AF / "paper" / "data" / "gold_dual_annotator.json",
    "活跃", "双人标注一致性数据")
add("axiom-finder/paper/", "Lane A", "data/distractors.json", AF / "paper" / "data" / "distractors.json",
    "活跃", "30 项人工构造的 distractor")
add("axiom-finder/paper/", "Lane A", "data/annotator_notes.md", AF / "paper" / "data" / "annotator_notes.md",
    "活跃", "标注者笔记")
add("axiom-finder/paper/", "Lane A", "data/build_gold.py", AF / "paper" / "data" / "build_gold.py",
    "活跃", "Gold 构建脚本")
add("axiom-finder/paper/", "Lane A", "data/fix_tier.py", AF / "paper" / "data" / "fix_tier.py",
    "活跃", "Tier 修复脚本")
add("axiom-finder/paper/", "Lane A", "data/fix_tier2.py", AF / "paper" / "data" / "fix_tier2.py",
    "活跃", "Tier 修复脚本 v2")
add("axiom-finder/paper/", "Lane A", "data/lane_c_stats.py", AF / "paper" / "data" / "lane_c_stats.py",
    "活跃", "Lane C 统计脚本")
add("axiom-finder/paper/", "Lane A", "data/rubric.md", AF / "paper" / "data" / "rubric.md",
    "活跃", "5-dim 评分 rubric 说明")
add("axiom-finder/paper/", "Lane A", "data/sanity_check.py", AF / "paper" / "data" / "sanity_check.py",
    "活跃", "Gold 健全性检查脚本")
add("axiom-finder/paper/", "Lane A", "data/sanity_check.txt", AF / "paper" / "data" / "sanity_check.txt",
    "活跃", "健全性检查输出")

add("axiom-finder/paper/", "Lane B 结果", "results/lane_b_predictions.json", AF / "paper" / "results" / "lane_b_predictions.json",
    "活跃", "Lane B 103 项预测(2026-06-25 recomposed,QWK 0.195)")
add("axiom-finder/paper/", "Lane B 结果", "results/lane_b_predictions.json.pre-recompose", AF / "paper" / "results" / "lane_b_predictions.json.pre-recompose",
    "备份", "recompose 前的备份(原 self-critique,QWK 0.171)")
add("axiom-finder/paper/", "Lane B 结果", "results/lane_b_predictions.api_key_missing.bak", AF / "paper" / "results" / "lane_b_predictions.api_key_missing.bak",
    "过时", "API key 缺失时的失败备份")
add("axiom-finder/paper/", "Lane B 结果", "results/lane_b_predictions.corrupted.json.bak", AF / "paper" / "results" / "lane_b_predictions.corrupted.json.bak",
    "过时", "损坏的备份文件")
add("axiom-finder/paper/", "Lane B 结果", "results/lane_c_report.md", AF / "paper" / "results" / "lane_c_report.md",
    "活跃", "Lane C 报告(MD)")
add("axiom-finder/paper/", "Lane B 结果", "results/lane_c_feedback.json", AF / "paper" / "results" / "lane_c_feedback.json",
    "活跃", "Lane C 反馈数据(用于 prompt 迭代)")
add("axiom-finder/paper/", "Lane B 结果", "results/lane_c_stats.json", AF / "paper" / "results" / "lane_c_stats.json",
    "活跃", "Lane C 统计数据(JSON)")

# formal/
add("axiom-finder/formal/", "Lean 4 (Phase A/D)", "AxiomForge.lean", AF / "formal" / "AxiomForge.lean",
    "活跃(阻塞)", "Lean 4 主入口,等 elan/lake 安装")
add("axiom-finder/formal/", "Lean 4", "lakefile.lean", AF / "formal" / "lakefile.lean",
    "活跃", "Lean 项目构建文件")
add("axiom-finder/formal/", "Lean 4", "lean-toolchain", AF / "formal" / "lean-toolchain",
    "活跃", "Lean 版本锁定")
add("axiom-finder/formal/", "Lean 4", "README.md", AF / "formal" / "README.md",
    "活跃", "Lean 集成说明")
add("axiom-finder/formal/", "Lean 4", "AxiomForge/", AF / "formal" / "AxiomForge",
    "活跃", "Lean module 目录")
add("axiom-finder/formal/", "Lean 4 examples", "examples/SC_counterexamples.lean", AF / "formal" / "examples" / "SC_counterexamples.lean",
    "活跃", "Structural Consistency 反例 Lean 示例")
add("axiom-finder/formal/", "Lean 4 examples", "examples/Thomson_examples.lean", AF / "formal" / "examples" / "Thomson_examples.lean",
    "活跃", "Thomson 投票定理 Lean 示例(Phase D)")
add("axiom-finder/formal/", "Lean 4 examples", "examples/Z3_examples.lean", AF / "formal" / "examples" / "Z3_examples.lean",
    "活跃", "Z3 验证的 Lean 示例")

# scripts/
add("axiom-finder/scripts/", "工具脚本", "recompose_lane_b.py", AF / "scripts" / "recompose_lane_b.py",
    "活跃", "用缓存的 layer1b/layer2_z3 重算 final_scores,不调 LLM")

# legacy_v0.2_pipeline/
add("axiom-finder/legacy_v0.2_pipeline/", "历史:v0.2", "README.md", AF / "legacy_v0.2_pipeline" / "README.md",
    "过时", "说明:v0.2 pipeline 已废弃,保留供历史参考")
add("axiom-finder/legacy_v0.2_pipeline/", "历史:v0.2", "pipeline.py", AF / "legacy_v0.2_pipeline" / "pipeline.py",
    "过时", "v0.2 5-agent 流水线编排(被 v0.3-alpha CLI 取代)")
add("axiom-finder/legacy_v0.2_pipeline/", "历史:v0.2", "axiom_deriver.py", AF / "legacy_v0.2_pipeline" / "axiom_deriver.py",
    "过时", "v0.2 axiom 派生器(被 KB nodes 取代)")
add("axiom-finder/legacy_v0.2_pipeline/", "历史:v0.2", "axiom_deriver_v2.py", AF / "legacy_v0.2_pipeline" / "axiom_deriver_v2.py",
    "过时", "v0.2 axiom 派生器 v2")
add("axiom-finder/legacy_v0.2_pipeline/", "历史:v0.2", "completeness_auditor.py", AF / "legacy_v0.2_pipeline" / "completeness_auditor.py",
    "过时", "v0.2 完整性审计")
add("axiom-finder/legacy_v0.2_pipeline/", "历史:v0.2", "completeness_rewriter.py", AF / "legacy_v0.2_pipeline" / "completeness_rewriter.py",
    "过时", "v0.2 自动重写")
add("axiom-finder/legacy_v0.2_pipeline/", "历史:v0.2", "consequence_predictor.py", AF / "legacy_v0.2_pipeline" / "consequence_predictor.py",
    "过时", "v0.2 后果预测")
add("axiom-finder/legacy_v0.2_pipeline/", "历史:v0.2", "diktat_aware_agents.py", AF / "legacy_v0.2_pipeline" / "diktat_aware_agents.py",
    "过时", "v0.2 diktat-aware agent")
add("axiom-finder/legacy_v0.2_pipeline/", "历史:v0.2", "literature_node_loader.py", AF / "legacy_v0.2_pipeline" / "literature_node_loader.py",
    "过时", "v0.2 literature loader")
add("axiom-finder/legacy_v0.2_pipeline/", "历史:v0.2", "memo_writer.py", AF / "legacy_v0.2_pipeline" / "memo_writer.py",
    "过时", "v0.2 memo writer")
add("axiom-finder/legacy_v0.2_pipeline/", "历史:v0.2", "notation_definer.py", AF / "legacy_v0.2_pipeline" / "notation_definer.py",
    "过时", "v0.2 notation definer")
add("axiom-finder/legacy_v0.2_pipeline/", "历史:v0.2", "perturbation_sampler.py", AF / "legacy_v0.2_pipeline" / "perturbation_sampler.py",
    "过时", "v0.2 perturbation sampler")
add("axiom-finder/legacy_v0.2_pipeline/", "历史:v0.2", "value_evaluator.py", AF / "legacy_v0.2_pipeline" / "value_evaluator.py",
    "过时", "v0.2 value evaluator")

# 第二步 + 第三步 — 2026-06-26 合并到 knowledge-base/ingest/
# (整目录已删;4 个 .py + z3_verify.py + 中文 docx/spec 已合并)
# 备份位置:git history (如果有) 或已清理;陛下应从 working tree 拉取
# 新位置:
#   knowledge-base/ingest/{discover, formalize, backtranslate, pipeline, z3_verify}.py
#   knowledge-base/SPEC_cn.md  (中文 spec)
#   knowledge-base/_archive/z3_verify_v0.py  (旧版 v0.3)

# training/
add("axiom-finder/training/", "训练", "diktat_injection.py", AF / "training" / "diktat_injection.py",
    "活跃", "diktat 注入训练脚本")
add("axiom-finder/training/", "训练", "few_shot/", AF / "training" / "few_shot",
    "活跃", "few-shot 训练样本目录")
add("axiom-finder/training/", "训练", "graph/", AF / "training" / "graph",
    "活跃", "训练图数据")
add("axiom-finder/training/", "训练", "seeds/", AF / "training" / "seeds",
    "活跃", "训练种子")
add("axiom-finder/training/", "训练", "structural_consistency/", AF / "training" / "structural_consistency",
    "活跃", "Structural Consistency axiom 训练数据")
add("axiom-finder/training/", "训练", "value_roots_draft.json", AF / "training" / "value_roots_draft.json",
    "活跃", "value_roots 草稿 JSON")

# outputs/
add("axiom-finder/outputs/", "v0.2 输出", "v0.2_shap/AX-STRUCTURAL-CONSISTENCY-001.md",
    AF / "outputs" / "v0.2_shap" / "AX-STRUCTURAL-CONSISTENCY-001.md",
    "历史", "v0.2 阶段产出的 SC axiom 描述(被 v0.3 取代)")
add("axiom-finder/outputs/", "v0.2 输出", "v0.2_shap/lundberg_2017_shap_PERT-STRUCTURAL_L17-CHAR-FN.json",
    AF / "outputs" / "v0.2_shap" / "lundberg_2017_shap_PERT-STRUCTURAL_L17-CHAR-FN.json",
    "历史", "v0.2 SHAP perturbation 结构化输出")
add("axiom-finder/outputs/", "v0.2 输出", "v0.2_shap/lundberg_2017_shap_PERT-STRUCTURAL_L17-CHAR-FN.md",
    AF / "outputs" / "v0.2_shap" / "lundberg_2017_shap_PERT-STRUCTURAL_L17-CHAR-FN.md",
    "历史", "v0.2 SHAP perturbation 报告(MD)")

# deploy/
add("axiom-finder/deploy/", "部署", "axiomfinder.pem", AF / "deploy" / "axiomfinder.pem",
    "活跃", "EC2 部署用的 SSH 私钥(注意不要提交)")
add("axiom-finder/deploy/", "部署", "README.md", AF / "deploy" / "README.md",
    "活跃", "部署说明")

# ── Important paper and methdology/ — 2026-06-26 陛下删除 ──
# (整目录已删;以下 11 PDF + 1 docx 历史记录保留作 archive)
# papers = [
#     ("Autoformalization of Game Descriptions...", "活跃", "..."),
#     ...
# ]

print(f"Total rows: {len(rows)}")
print(f"Statuses: {dict(Counter(r['status'] for r in rows))}")

# ── Write xlsx ──
wb = openpyxl.Workbook()

# Define styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill_active = PatternFill("solid", fgColor="2E7D32")      # 活跃 green
header_fill_history = PatternFill("solid", fgColor="FFA726")     # 历史 orange
header_fill_obsolete = PatternFill("solid", fgColor="9E9E9E")    # 过时 gray
header_fill_indep = PatternFill("solid", fgColor="1976D2")       # 独立 blue
header_fill_temp = PatternFill("solid", fgColor="EF5350")        # 临时 red
header_fill_backup = PatternFill("solid", fgColor="AB47BC")      # 备份 purple
header_fill_blocked = PatternFill("solid", fgColor="FF6F00")     # 阻塞 dark orange

status_fill = {
    "活跃": header_fill_active,
    "活跃(阻塞)": header_fill_blocked,
    "历史": header_fill_history,
    "历史/参考": header_fill_history,
    "过时": header_fill_obsolete,
    "独立项目": header_fill_indep,
    "临时": header_fill_temp,
    "备份": header_fill_backup,
    "参考": header_fill_indep,
}

thin = Side(border_style="thin", color="BDBDBD")
cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Sheet 1: All files
ws = wb.active
ws.title = "全部文件清单"

headers = ["顶层区域", "子项目/阶段", "文件名", "相对路径", "大小", "修改日期", "状态", "作用说明"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill_active
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = cell_border

for i, r in enumerate(rows, 2):
    vals = [r["zone"], r["project"], r["file"], r["rel_path"],
            r["size"], r["mtime"], r["status"], r["purpose"]]
    fill = status_fill.get(r["status"], None)
    for col, v in enumerate(vals, 1):
        c = ws.cell(row=i, column=col, value=v)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = cell_border
        if fill and col == 7:  # status column colored
            c.fill = fill
            c.font = Font(bold=True, color="FFFFFF")

# Column widths
widths = [22, 18, 50, 45, 10, 12, 14, 65]
for col, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = w

ws.row_dimensions[1].height = 28
ws.freeze_panes = "A2"

# Sheet 2: 仅过时/历史/临时
ws2 = wb.create_sheet("仅过时+历史+临时")
for col, h in enumerate(headers, 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill_obsolete
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = cell_border

obsolete = [r for r in rows if r["status"] in ("过时", "历史", "临时", "备份", "参考", "历史/参考")]
for i, r in enumerate(obsolete, 2):
    vals = [r["zone"], r["project"], r["file"], r["rel_path"],
            r["size"], r["mtime"], r["status"], r["purpose"]]
    fill = status_fill.get(r["status"], header_fill_obsolete)
    for col, v in enumerate(vals, 1):
        c = ws2.cell(row=i, column=col, value=v)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = cell_border
        if col == 7:
            c.fill = fill
            c.font = Font(bold=True, color="FFFFFF")

for col, w in enumerate(widths, 1):
    ws2.column_dimensions[get_column_letter(col)].width = w
ws2.row_dimensions[1].height = 28
ws2.freeze_panes = "A2"

# Sheet 3: 项目阶段图谱
ws3 = wb.create_sheet("项目阶段图谱")
ws3.column_dimensions["A"].width = 14
ws3.column_dimensions["B"].width = 60
ws3.column_dimensions["C"].width = 80

ws3.cell(row=1, column=1, value="阶段").font = header_font
ws3.cell(row=1, column=2, value="目标 / 交付物").font = header_font
ws3.cell(row=1, column=3, value="涉及文件 / 目录").font = header_font
for col in (1, 2, 3):
    ws3.cell(row=1, column=col).fill = header_fill_active
    ws3.cell(row=1, column=col).alignment = Alignment(horizontal="center")
    ws3.cell(row=1, column=col).border = cell_border

phases = [
    ("Phase 0", "项目脚手架 + 安全清理",
     "axiom-finder/ 根目录(README/LICENSE/requirements/SHIP_TO_GITHUB.sh)\n.gitignore + CI workflow"),
    ("Phase 1", "Lane A: 构建黄金标准",
     "paper/data/gold.json (104项)\npaper/data/gold_dual_annotator.json\npaper/data/distractors.json\npaper/data/build_gold.py + sanity_check.py\nPLAN/02_lane_A_gold_standard.docx"),
    ("Phase 2", "Lane B: 5-dim rubric + 3-Layer 评估",
     "knowledge-base/ingest/lane_b_evaluator.py\nkb/ingest/lane_b_prompts/v3.md\nkb/ingest/z3_verify.py (Layer 2)\npaper/results/lane_b_predictions.json\nscripts/recompose_lane_b.py (迭代工具)\nPLAN/03_lane_B_evaluator.md"),
    ("Phase 3", "Lane C: 校准统计 + 反馈循环",
     "paper/data/lane_c_stats.py\npaper/results/lane_c_report.md\npaper/results/lane_c_stats.json\npaper/results/lane_c_feedback.json\nPLAN/04_lane_C_statistics.md"),
    ("Phase 4", "Lane D: 论文 + 形式化",
     "PLAN/05_lane_D_paper.md\nformal/AxiomForge.lean\nformal/lakefile.lean + lean-toolchain\nformal/examples/ (SC/Thomson/Z3 Lean 示例)\n第三步/PIPELINE_SPEC.md"),
    ("Phase 2 Pipeline", "v0.3 候选发现 D⇄F 交错 + Z3 验证,写 AxiomRecord 到 JSONL",
     "第二步/{discover, formalize, backtranslate, pipeline}.py\n第三步/PIPELINE_SPEC.md (D⇄F 交错规格)\n第三步/z3_verify.py (Tier A/B/C reference)\n第三步/axiom_forge_step3_cn.docx + 第二步/axiom_forge_step2_cn.docx (中文 spec)"),
    ("Phase A", "Lean 4 形式化(阻塞:等陛下装 elan)",
     "formal/AxiomForge.lean\nformal/examples/*.lean\nPLAN/09_lean_integration.md\n部署:curl elan-init.sh | sh + lake exe cache get"),
    ("Phase B", "文献 ingest(已落地 57 papers + 92 drafts)",
     "knowledge-base/ingest/literature_fetcher.py\nkb/ingest/extract_nodes.py\nkb/ingest/paper_pipeline.py\nkb/ingest/extract_prompts/ + build_seeds.py\nagents/literature_agent.py"),
    ("Phase C", "Proof-checker agent(Lean 集成)",
     "knowledge-base/ingest/proof_checker.py"),
    ("Phase D", "Thomson 2023 形式化(4-8 周)",
      "formal/examples/Thomson_examples.lean"),
    ("Phase E", "政治哲学/投票/历史(2-3 月,待启动)",
     "knowledge-base/nodes/value_anchors/ (33 节点)\nagents/gap_finder.py (候选 axiom 发现)"),
    ("KB 节点", "85 节点知识库",
     "knowledge-base/nodes/{axioms, theorems, assumptions, value_anchors, diktats, tradeoffs, scenarios, relations, literature, compressed_views, training_signals}\nkb/SCHEMA.md + UPDATE_MECHANISM.md"),
    ("规划文档", "9 个 .docx + 7 个 .md 计划",
     "PLAN/01-12 (master/各 Lane/engineering/git/lean/glossary/checklist/risk)"),
    ("Web 部署", "FastAPI + SPA UI + EC2",
     "web_api.py + static/\ndeploy/axiomfinder.pem + deploy/README.md"),
    ("历史归档", "v0.2 5-agent pipeline",
     "legacy_v0.2_pipeline/ (13 文件,标注过时,保留参考)\noutputs/v0.2_shap/ (3 文件,标注历史)"),
    ("独立项目", "与 axiom-forge 无关,但占用 Downloads",
     "'research applied'/ (信贷市场 + SHAP 实验,陛下独立研究)\n~$*.docx (10 个 Word 临时锁)\n2026-06-26 已清理:output/(个人破产爬虫)+ Important paper and methdology/(11 PDF)整目录删除"),
]

for i, (ph, goal, files) in enumerate(phases, 2):
    c1 = ws3.cell(row=i, column=1, value=ph)
    c1.font = Font(bold=True)
    c1.alignment = Alignment(vertical="top", wrap_text=True)
    c1.border = cell_border
    c1.fill = PatternFill("solid", fgColor="E3F2FD")

    c2 = ws3.cell(row=i, column=2, value=goal)
    c2.alignment = Alignment(vertical="top", wrap_text=True)
    c2.border = cell_border

    c3 = ws3.cell(row=i, column=3, value=files)
    c3.alignment = Alignment(vertical="top", wrap_text=True)
    c3.border = cell_border

ws3.row_dimensions[1].height = 28
ws3.freeze_panes = "A2"

# Save
out_path = DL / "downloads_inventory.xlsx"
wb.save(out_path)
print(f"\nSaved: {out_path}")
print(f"Size: {fmt_size(out_path)}")
print(f"\nSheet 1 '全部文件清单': {len(rows)} 行")
print(f"Sheet 2 '仅过时+历史+临时': {len(obsolete)} 行")
print(f"Sheet 3 '项目阶段图谱': {len(phases)} 个阶段")